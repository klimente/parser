import requests
from bs4 import BeautifulSoup
import re
import math
import openpyxl

def get_url(url):
    r = requests.get(url)
    return r.text                 #возвращаем html код страницы от url , полученного в точке входа


def get_html(html):                                 #Функция для взятия матча с данного урла , откуда достаем две ссылки на команды в матче соответственно
    soup = BeautifulSoup(html, 'lxml')
    everymatch = soup.find('div',id='block-system-main').find_all('div',class_='black-title-content-border')
    h=[]

    for k in everymatch:
        c=k.find_all('div',class_ = 'match')
        for i in c:
            home = i.find('td',class_='col-2').find('div', class_='home')
            visit =  i.find('td',class_='col-2').find('div', class_='visit')
        #i.find('a').get('href')
            o = []
            o.append( home.find('a').get('href'))
            o.append(visit.find('a').get('href'))
            h.append(o)

    print(h)
    return h



def get_raspisaniya(html):                          #Ф-ция , которая возвращает результаты матча внутри страницы команды
    regexp = r'^/teams/teamfixtures/'               #Использую регулярные выражения для поиска ссылок , потому что список без классов
    soup = BeautifulSoup(html,'lxml')
    rasp = soup.find('div',class_='entity-menu').find('a',href=re.compile(regexp)).get('href')
    return rasp

def get_results_h(html):                              #Ф-ция , в которой достаю все необходиые данные для анализа игры
    soup = BeautifulSoup(html, 'lxml')
    title = soup.find('title').text
    chtitle = str(title).split()[0]
    score = soup.find('div',id='block-system-main').find('div',class_='block').find_all('div',class_='match')
    amount = 0
    games_at_home = 0
    home_goals = 0
    miss_goals_from_v = 0
    #print(chtitle)
    games_visit = 0
    visit_goals =0
    miss_goals_from_h = 0
    for block in score:
        home = block.find('td',class_='col-2').find('div',class_='home').find('td',class_='team-name').find('a').string

        homescore = block.find('td',class_='col-2').find('div',class_='home').find('td',class_='sc').find('span').string.split()[0]

        if home.split()[0] == chtitle:
            if homescore == '-':
                homescore = 0
            games_at_home+=1
            home_goals += float(homescore)
        else:
            if homescore == '-':
                homescore = 0
            miss_goals_from_h += float(homescore)

        visit = block.find('td',class_='col-2').find('div',class_='visit').find('td',class_='team-name').find('a').string

        visitscore = block.find('td',class_='col-2').find('div',class_='visit').find('td',class_='sc').find('span').string.split()[0]

        if visit.split()[0] == chtitle:
            if visitscore == '-':
                visitscore = 0
            games_visit+=1
            visit_goals += int(visitscore)
            #print(visit_goals,'-gost-',visitscore)
        else:
            if visitscore == '-':
                visitscore = 0
            miss_goals_from_v += int(visitscore)

        #print(home,' ',visit)
        #print(homescore,' ',visitscore)
        amount+=1

    average_goals_h= home_goals/games_at_home
    average_goals_missed = miss_goals_from_v/games_at_home
    return games_at_home,games_visit,home_goals,visit_goals, average_goals_h,average_goals_missed,chtitle

def get_results_v(html):                              #Ф-ция , в которой достаю все необходиые данные для анализа игры
    average_goals_v = 0
    soup = BeautifulSoup(html, 'lxml')
    title = soup.find('title').text
    chtitle = str(title).split()[0]
    score = soup.find('div',id='block-system-main').find('div',class_='block').find_all('div',class_='match')
    amount = 0
    games_at_home = 0
    home_goals = 0
    miss_goals_from_v = 0
    #print(chtitle)
    games_visit = 0
    visit_goals =0
    miss_goals_from_h = 0
    for block in score:
        home = block.find('td',class_='col-2').find('div',class_='home').find('td',class_='team-name').find('a').string

        homescore = block.find('td',class_='col-2').find('div',class_='home').find('td',class_='sc').find('span').string.split()[0]

        if home.split()[0] == chtitle:
            if homescore == '-':
                homescore = 0
            games_at_home+=1
            home_goals += float(homescore)
        else:
            if homescore == '-':
                homescore = 0
            miss_goals_from_h += float(homescore)

        visit = block.find('td',class_='col-2').find('div',class_='visit').find('td',class_='team-name').find('a').string

        visitscore = block.find('td',class_='col-2').find('div',class_='visit').find('td',class_='sc').find('span').string.split()[0]

        if visit.split()[0] == chtitle:
            if visitscore == '-':
                visitscore = 0
            games_visit+=1
            visit_goals += int(visitscore)
            #print(visit_goals,'-gost-',visitscore)
        else:
            if visitscore == '-':
                visitscore = 0
            miss_goals_from_v += int(visitscore)

        #print(home,' ',visit)
        #print(homescore,' ',visitscore)
        amount+=1
    if games_visit != 0:

        average_goals_v= float(visit_goals/games_visit)
    average_goals_missed_v = float(miss_goals_from_v/games_visit)
    return games_at_home, games_visit,home_goals,visit_goals,average_goals_v,average_goals_missed_v,chtitle




def calculation(total_game_h,total_goals_h,total_game_v,total_goals_v,averege_t1_attack,average_goals_missed,averege_goals_v,averege_missed_v):
    average_number_scored_home = float(total_goals_h/total_game_h)
    average_number_scored_visit = float(total_goals_v/total_game_v)
    attack_strength_home = float(averege_t1_attack/average_number_scored_home)
    defensive_strength_visit = float(average_goals_missed/average_number_scored_home)
    math_expectation_t1 = float(attack_strength_home*defensive_strength_visit*average_number_scored_home)
    attack_strength_visit = float(averege_goals_v/average_number_scored_visit)
    defensive_strength_home = float(averege_missed_v/average_number_scored_visit)
    math_expectation_t2 = float(attack_strength_visit*defensive_strength_home*average_number_scored_visit)



    return math_expectation_t1,math_expectation_t2

def poisson_probability(m_e_h,m_e_v):
    p = []
    for i in range(6):
        p.append(((m_e_h**i)*math.exp(-m_e_h))/math.factorial(i))
    #print(p)
    prob = []
    for i in range(6):
        prob.append(((m_e_v**i)*math.exp(-m_e_v))/math.factorial(i))
    #print(prob)
    big_one =[]
    probabilt_total_m = 0
    for i in p:
        for x in prob:
            big_one.append(i*x)

            if (p.index(i) <3) and (prob.index(x) <3):
                probabilt_total_m += i*x
                #print(p.index(i), prob.index(x), i * x,probabilt_total_m)
    return probabilt_total_m


def main():
    url = 'https://www.soccer0010.com/matches-list'
    teams = get_html(get_url(url))
    #print(teams)

    for item in teams:
        url_or = 'https://www.soccer0010.com'
        match_total_game_h = 0
        match_total_goals_h = 0
        match_total_game_v = 0
        match_total_goals_v = 0
        average_goals_h = 0.0
        average_goals_missed = 0.0
        average_goals_v = 0.0
        average_goals_missed_v = 0.0
        match_name = []
        for t in item:
            url_for_team = url_or + t
            #print(url_for_team)

            url_for_result =url_or + get_raspisaniya(get_url(url_for_team))
      #  match_total_goal +=

            print(item.index(t))
            if item.index(t) == 0:
                try:
                    match_total_game_h += get_results_h(get_url(url_for_result))[0]
                    match_total_goals_h += get_results_h(get_url(url_for_result))[2]
                    match_total_game_v += get_results_h(get_url(url_for_result))[1]
                    match_total_goals_v += get_results_h(get_url(url_for_result))[3]
                    average_goals_h = get_results_h(get_url(url_for_result))[4]
                    average_goals_missed = get_results_h(get_url(url_for_result))[5]
                    match_name.append( get_results_h(get_url(url_for_result))[6])
                except ZeroDivisionError:
                    print('exept')
                    match_total_game_h = 0
                    match_total_goals_h = 0
                    match_total_game_v = 0
                    match_total_goals_v = 0
                    average_goals_h = 0
                    average_goals_missed = 0
            else:

                try:
                    match_total_game_h  += get_results_v(get_url(url_for_result))[0]
                    match_total_goals_h += get_results_v(get_url(url_for_result))[2]
                    match_total_game_v += get_results_v(get_url(url_for_result))[1]
                    match_total_goals_v += get_results_v(get_url(url_for_result))[3]
                    average_goals_v = get_results_v(get_url(url_for_result))[4]
                    average_goals_missed_v = get_results_v(get_url(url_for_result))[5]
                    match_name.append(get_results_v(get_url(url_for_result))[6])
                except ZeroDivisionError:
                    print('exept')
                    match_total_game_h += 0
                    match_total_goals_h += 0
                    match_total_game_v += 0
                    match_total_goals_v += 0
                    average_goals_v = 0
                    average_goals_missed_v = 0
            print(match_total_game_h,match_total_goals_h,average_goals_h,average_goals_missed_v,average_goals_v)
            if match_total_game_h == 0 or match_total_goals_h ==0 or match_total_game_v == 0 or match_total_goals_v == 0 or average_goals_h == 0.0 or average_goals_missed == 0.0 or average_goals_v == 0.0 or average_goals_missed_v == 0.0:
                continue
        try:
            print(calculation(match_total_game_h,match_total_goals_h,match_total_game_v,match_total_goals_v,average_goals_h,average_goals_missed,average_goals_v,average_goals_missed_v))
        except  ZeroDivisionError:
            print('exept')
            continue
        if calculation(match_total_game_h,match_total_goals_h,match_total_game_v,match_total_goals_v,average_goals_h,average_goals_missed,average_goals_v,average_goals_missed_v)[0] == 0:
            m_e_h = 0.5
        else: m_e_h = calculation(match_total_game_h,match_total_goals_h,match_total_game_v,match_total_goals_v,average_goals_h,average_goals_missed,average_goals_v,average_goals_missed_v)[0]
        if calculation(match_total_game_h,match_total_goals_h,match_total_game_v,match_total_goals_v,average_goals_h,average_goals_missed,average_goals_v,average_goals_missed_v)[1] == 0.0:
            m_e_v = 0.5
        else:
            m_e_v = calculation(match_total_game_h,match_total_goals_h,match_total_game_v,match_total_goals_v,average_goals_h,average_goals_missed,average_goals_v,average_goals_missed_v)[1]




        total_m = poisson_probability(m_e_h,m_e_v)

        #print(total_m)
        try:
            if total_m < 0.7 :
                wb = openpyxl.load_workbook(filename='test.xlsx')
                sheet = wb['test']
                sheet.cell(row = teams.index(item),column=2).value = match_name[0]
                sheet.cell(row=teams.index(item), column=3).value = match_name[1]
                sheet.cell(row=teams.index(item), column=4).value = total_m
                print(match_name,'-',total_m)
                wb.save('test.xlsx')
        except IndexError:
            continue


if __name__ == '__main__':
    main()